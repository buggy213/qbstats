import wget
import os
from openpyxl import load_workbook
from collections import OrderedDict
from itertools import islice, takewhile, repeat

# SQBS File Format
# https://www.qbwiki.com/wiki/SQBS_data_file
# sqbs_file_format

# Checks if SQBS file exists; if it does, overwrite
# Writes team names and team members to the file

# TODO: Check for cases where room is on next round before TUH = 20 in previous round

def write_rosters(filename, rosters):
    if os.path.isfile(filename):
        os.remove(filename)

    f = open(filename, 'w+')

    write_line(f, len(rosters))
    for team_name, members in rosters.items():
        write_line(f, len(members) + 1)
        write_line(f, team_name)
        for member in members:
            write_line(f, member)

    return f

def write_match(file, match, id, rosters, round, side = None):
    write_line(file, id)
    left_team_name = match[0][1].strip()
    right_team_name = match[0][7].strip()
    left_team_index = list(rosters.keys()).index(left_team_name)
    right_team_index = list(rosters.keys()).index(right_team_name)

    tuh = 0

    if side is None:
        left_team_score = get_value(1, 1, match)
        right_team_score = get_value(1, 7, match)
        tuh = get_value(8, 0, match) - 1
    elif side is True:
        left_team_score = -1
        right_team_score = 0
    elif side is False:
        left_team_score = 0
        right_team_score = -1
    
    left_team_bonuses_score = get_value(1, 0, match)
    right_team_bonuses_score = get_value(2, 0, match)

    left_players = list()
    right_players = list()

    left_bonuses_heard = 0
    right_bonuses_heard = 0

    # Hopefully the word player isn't an actual name LOL
    for column in range(1, 7):
        if "Player" in match[2][column]:
            left_players.append(empty_player())
        else:
            player = read_player(column, match, rosters, left_team_name, tuh)
            left_bonuses_heard += player[2] + player[3] # powers + tens
            left_players.append(player)
    for column in range(7, 13):
        if "Player" in match[2][column]:
            right_players.append(empty_player())
        else:
            player = read_player(column, match, rosters, right_team_name, tuh)
            right_bonuses_heard += player[2] + player[3]
            right_players.append(player)

    # Pad out to 16 players because for some fucking reason that's what SQBS expects
    left_players.append(empty_player())
    left_players.append(empty_player())
    right_players.append(empty_player())
    right_players.append(empty_player())

    write_lines(file, id, left_team_index, right_team_index, left_team_score, right_team_score, \
        tuh, round, left_bonuses_heard, left_team_bonuses_score, right_bonuses_heard, right_team_bonuses_score, \
        '0', '0', '0', \
        '0', '0')

    for i in range(8):
        write_player(file, left_players[i])
        write_player(file, right_players[i])

def empty_player():
    return '-1', '0', '0', '0', '0', '0'

def read_player(column, cells, rosters, team_name, tuh):
    index = rosters[team_name].index(cells[2][column].strip())
    gp = get_value(3, column, cells) / tuh
    powers = get_value(4, column, cells)
    tens = get_value(5, column, cells)
    negs = get_value(6, column, cells)
    total = get_value(7, column, cells)

    return index, gp, powers, tens, negs, total

def write_player(file, player):
    write_lines(file, player[0], player[1], player[2], player[3], player[4], '0', player[5])
    

def write_lines(file, *lines):
    for line in lines:
        write_line(file, str(line))

def get_value(y, x, cells):
    if isinstance(cells[y][x], str) and ':' in cells[y][x]:
        return int(cells[y][x].split(':')[1])
    else:
        return int(cells[y][x])

# Helper function to write a string followed by a newline
# (why isn't this a builtin?!?!?)
def write_line(f, content):
    f.write(str(content) + '\n')


def split_every(n, iterable):
    """
    Slice an iterable into chunks of n elements
    :type n: int
    :type iterable: Iterable
    :rtype: Iterator
    """
    iterator = iter(iterable)
    return takewhile(bool, (list(islice(iterator, n)) for _ in repeat(None)))


def main():

    tournament_name = "NeATo" # input("Tournament name: ")

    # while True:
    #     url = input("URL of published aggregate spreadsheet: ")
    #     try:
    #         wget.download(url)
    #     except ValueError:
    #         print("Invalid URL, try again")
    #         continue
    #     break
    
    xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    if len(xlsx_files) != 1:
        raise ValueError('should be only one xlsx file in the current directory')

    spreadsheet = xlsx_files[0]

    workbook = load_workbook(filename=spreadsheet, data_only=True)
    
    # Support for double division tournaments
    teams = OrderedDict()

    rosters_worksheet = workbook["Rosters"]
    for row in rosters_worksheet.values:
        if row[0] is None:
            continue
        team = row[0].strip()
        members = tuple(member for member in row[1:] if member is not None)
        teams[team] = members
    
    file_descriptor = write_rosters("sqbs", teams)
    num_matches_index = len(file_descriptor.readlines())
    write_line(file_descriptor, '------')
    id = 0
    for worksheet_name in workbook.sheetnames:
        if worksheet_name == "Rosters":
            continue

        worksheet = workbook[worksheet_name]
        round = worksheet_name.split(' ')[1]
        for match in split_every(10, worksheet.values):
            if match[0][0] is None:
                continue
            if int(match[8][0].split(':')[1]) != 1:
                write_match(file_descriptor, match, id, teams, round)
            elif match[0][1] != "Team A":
                # Dealing with forfeits: prompt for each match that could be forfeit
                print("Potential forfeit (y to accept) -- Round " + round + " between " + match[0][1] + " and " + match[0][7])
                forfeit = 'y' in input("y/n")
                if forfeit:
                    side = input("Which team forfeited? (A|B): ")
                    write_match(file_descriptor, match, id, teams, round, side is 'A')
                else:
                    continue
            else:
                continue

            id += 1

    contents = file_descriptor.readlines()
    contents.insert(num_matches_index, str(id + 1))
    file_descriptor.truncate(0)
    file_descriptor.writelines(contents)

    write_lines(
        file_descriptor, '1', '1', '3', '0', '1', '1', '0', '1', '1', \
        '1', '1', '1', '1', '0', '0', '0', '1', tournament_name, '', '', '', '', \
        '0', '_rounds.html', '_standings.html', '_individuals.html', '_games.html', '_teamdetail.html', '_playerdetail.html', '_statkey.html', \
        '', '0', str(len(teams)), '-1', \
        '15', '10', '-5', '0', str(len(teams)))

    for team_name, members in teams.items():
        write_line(file_descriptor, '0')

    file_descriptor.close()

if __name__ == '__main__':
    main()